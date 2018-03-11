import openpyxl
import time


START = time.time() #start counting...


'''
eating data
'''
file_tun = openpyxl.load_workbook('''your path here''')
file_active = openpyxl.load_workbook('''your path here''')
oper_tun = file_tun.get_sheet_by_name('Summary') #a single cell of tun protein
oper_active = file_active.get_sheet_by_name('Summary') #a single cell of active protein


'''
creating destination
'''
dest = openpyxl.Workbook()
sh1 = dest.active
sh1.title = 'emPAI calculations'
sh1['A1'] = 'Active accession num'
sh1['D1'] = 'Tun name'
sh1['B1'] = 'Tun accession num'
sh1['E1'] = 'Active emPAI'
sh1['C1'] = 'Active name'
sh1['F1'] = 'Tun emPAI'
sh1['G1'] = 'Active - Tun'
sh1['H1'] = 'Tun - Active'
sh1['I1'] = 'Tun/Active ratio'


'''
defining list constructors
'''
tun_list = []
active_list = []
tun_list_copy = []
active_list_copy = []
temp = []
tun_sorted = []
active_sorted = []


'''
testing if it works xD
'''
print('TUN: ', oper_tun['D10'].value)
print('ACTIVE: ', oper_active['D10'].value)


'''
parsing tun table
'''
for i in range(6, 3225):
    acc = 'D' + str(i)
    nam = 'E' + str(i)
    emp = 'M' + str(i)
    seq = 'L' + str(i)
    if oper_tun[acc].value != None:
        #print('Not none tun: ', type(oper_tun[emp].value))
        temp.append(oper_tun[acc].value)
        temp.append(oper_tun[nam].value)
        temp.append(oper_tun[emp].value)
        temp.append(oper_tun[seq].value)
        temp = tuple(temp)
        tun_list.append(temp) #adding to list
        temp = list(temp)
        temp.clear()
    elif oper_tun[acc].value == None:
        #print('none acc tun: ', type(oper_tun[acc].value))
        temp.append('NONE')
        temp.append(oper_tun[nam].value)
        temp.append(oper_tun[emp].value)
        temp.append(oper_tun[seq].value)
        temp = tuple(temp)
        tun_list.append(temp) #adding to list
        temp = list(temp)
        temp.clear()

print('tun_list len: ', len(tun_list))


'''
parsing active table
'''
for i in range(6, 3290):
    acc = 'D' + str(i)
    nam = 'E' + str(i)
    emp = 'M' + str(i)
    seq = 'L' + str(i)
    if oper_active[acc].value != None:
        #print('Not none active: ', type(oper_active[emp].value))
        temp.append(oper_active[acc].value)
        temp.append(oper_active[nam].value)
        temp.append(oper_active[emp].value)
        temp.append(oper_active[seq].value)
        temp = tuple(temp)
        active_list.append(temp) #adding to list
        temp = list(temp)
        temp.clear()
    elif oper_active[acc].value == None:
        #print('none acc active: ', type(oper_active[acc].value))
        temp.append('NONE')
        temp.append(oper_active[nam].value)
        temp.append(oper_active[emp].value)
        temp.append(oper_active[seq].value)
        temp = tuple(temp)
        active_list.append(temp) #adding to list
        temp = list(temp)
        temp.clear()

#print(active_list)
print('active_list len: ', len(active_list))
print('tun_list len: ', len(tun_list))

STOP = time.time()
elapsed = STOP - START
print(elapsed)
print('WHAT NOW?!')

'''
#doing matching
'''
k = 0
l = 0
for i in range(len(tun_list)):
    a = 'A' + str(k+2)
    b = 'B' + str(k+2)
    c = 'C' + str(k+2)
    d = 'D' + str(k+2)
    e = 'E' + str(k+2)
    f = 'F' + str(k+2)
    g = 'G' + str(k+2)
    h = 'H' + str(k+2)
    whyy = 'I' + str(k+2)
    for j in range(len(active_list)):
        if active_list[j][1] == tun_list[i][1] and active_list[j][0] == tun_list[i][0]:
            #accession
            sh1[a] = active_list[j][0]
            sh1[b] = tun_list[i][0]
            #name
            sh1[c] = active_list[j][1]
            sh1[d] = tun_list[i][1]
            #emPAI
            sh1[e] = active_list[j][2]
            sh1[f] = tun_list[i][2]
            #calc
            if active_list[j][2] != None and tun_list[i][2] != None:
                sh1[g] = (active_list[j][2] - tun_list[i][2])
                sh1[h] = (tun_list[i][2] - active_list[j][2])
                sh1[whyy] = (tun_list[i][2] / active_list[j][2])
            else:
                sh1[g] = 'NaN'
                sh1[h] = 'NaN'
                sh1[whyy] = 'NaN'
            k += 1
        elif active_list[j][1] == tun_list[i][1] and active_list[j][0] == 'NONE': #and active_list[j][1] == tun_list[i][1]:
            #print('no cos sie dzieje!~', l)
            #l += 1
            sh1[a] = 'None'
            sh1[b] = 'None'
            sh1[c] = active_list[j][1]
            sh1[d] = tun_list[j][1]
            k += 1


k = 0
for i in range(len(active_list)):
    if (tun_list.__contains__(active_list[i]) == False):
        row = str(3221+k)
        ind1 = 'A' + row
        ind2 = 'B' + row
        ind3 = 'C' + row
        ind4 = 'D' + row
        ind5 = 'E' + row
        ind6 = 'F' + row
        ind7 = 'G' + row
        ind8 = 'H' + row
        sh1[ind1] = active_list[i][0]
        sh1[ind2] = 'No match'
        sh1[ind3] = active_list[i][1]
        sh1[ind4] = 'No match'
        sh1[ind5] = active_list[i][2]
        sh1[ind6] = 'No match'
        sh1[ind7] = 'NaN'
        sh1[ind8] = 'NaN'        
        k += 1
    else:
        pass

dest.save(filename = '''destination''')